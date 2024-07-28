from openpyxl import load_workbook
from tkinter import messagebox
from tabulate import tabulate
import tkinter as tk
import pandas as pd
import time

def read_and_process_input_file(input_file):
    """Read and process the input file, extracting and filtering data."""
    df = pd.read_excel(input_file)
    headers = df.columns.tolist()
    og_data = [headers] + df.values.tolist()
    
    # Converting the columns into lists for more readable computation
    columns_list = [df[col].tolist() for col in df.columns]
    for col in columns_list:
        col.reverse()

    # Creating necessary filters and data
    indices_to_delete = {1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13, 14}
    filtered_headers = [header for idx, header in enumerate(headers) if idx not in indices_to_delete]
    filtered_headers.extend(['DAILY GAIN/LOSS', 'MARGIN', 'MARGIN CALL'])
    filtered_columns_list = [col for idx, col in enumerate(columns_list) if idx not in indices_to_delete]
    
    # Transposing lists of column data which was stored as rows, now to columns
    transposed_columns_list = list(zip(*filtered_columns_list))
    transposed_columns_list_with_headers = [filtered_headers] + transposed_columns_list
    
    return og_data, transposed_columns_list_with_headers

def write_to_sheet(wb, sheet_name, data):
    """Write data to the specified sheet, clearing it first if it exists."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    for r_idx, row in enumerate(data):
        for c_idx, value in enumerate(row):
            ws.cell(row=r_idx+1, column=c_idx+1, value=value)

def get_user_input():
    """Display a Tkinter dialog to get user input for various variables."""
    def on_submit():
        variables['VaR_Margin'] = float(entry_var_margin.get())/100
        variables['applicable_margin_rate'] = float(entry_applicable_margin_rate.get())/100
        variables['lot_size'] = int(entry_lot_size.get())
        variables['settle_price'] = float(entry_settle_price.get())
        variables['link'] = entry_link.get()

        if not all(variables.values()):
            messagebox.showwarning("Input Error", "All fields must be filled out.")
        
        root.destroy()

    variables = {}
    root = tk.Tk()
    root.title("Futures Data")

    tk.Label(root, text="VaR Margin | Enter % In Value:").pack(pady=5)
    entry_var_margin = tk.Entry(root)
    entry_var_margin.pack(pady=5)

    tk.Label(root, text="Applicable Margin Rate | Enter % In Value:").pack(pady=5)
    entry_applicable_margin_rate = tk.Entry(root)
    entry_applicable_margin_rate.pack(pady=5)

    tk.Label(root, text="Lot Size:").pack(pady=5)
    entry_lot_size = tk.Entry(root)
    entry_lot_size.pack(pady=5)

    tk.Label(root, text="Settle Price:").pack(pady=5)
    entry_settle_price = tk.Entry(root)
    entry_settle_price.pack(pady=5)

    tk.Label(root, text="Data Reference Link | Add N/A if not applicable").pack(pady=5)
    entry_link = tk.Entry(root)
    entry_link.pack(pady=5)

    submit_button = tk.Button(root, text="Submit", command=on_submit)
    submit_button.pack(pady=10)

    root.mainloop()

    return variables

def setup_input_file():
    """Set up the input file and write processed data to output files."""
    input_file = "Data.xlsx"
    output_file = "FSD_Input.xlsx"
    sheet_name_computing = "Computing Data"
    sheet_name_futures = "Futures Data"
    sheet_name_variables = "Variables"

    og_data, transposed_columns_list_with_headers = read_and_process_input_file(input_file)
    
    user_input = get_user_input()
    variables_data = [
        ['VaR_Margin', user_input['VaR_Margin']],
        ['Applicable_Margin_Rate', user_input['applicable_margin_rate']],
        ['Lot_Size', user_input['lot_size']],
        ['Settle_Price', user_input['settle_price']],
        ['Link', user_input['link']]
    ]

    wb = load_workbook(output_file)
    
    write_to_sheet(wb, sheet_name_computing, transposed_columns_list_with_headers)
    write_to_sheet(wb, sheet_name_futures, og_data)
    write_to_sheet(wb, sheet_name_variables, variables_data)
    
    wb.save(output_file)

def merge_workbooks(file1, file2, output_file, sheet_name1='Long Position', sheet_name2='Short Position'):
    """Merge two Excel workbooks into one with specified sheet names."""
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name=sheet_name1, index=False)
        df2.to_excel(writer, sheet_name=sheet_name2, index=False)

def compute_position(data, position, margin, lot_size, is_long=True):
    """Compute and append data for long or short positions."""
    data[1].append('0')
    data[1].append(margin)
    data[1].append('-')

    # Skipping header row and first row as no computation can be done there
    for i in data:
        if position < 2:
            position += 1
            continue
        
        # Extracting necessary data for positional computation
        previous_day = data[position - 1]
        current_day = data[position]

        # Computing contract gain or loss and appending the same
        gain_or_loss = round(float(lot_size * (current_day[1] - previous_day[1])), 2)
        gain_or_loss = gain_or_loss if is_long else -gain_or_loss
        i.append(gain_or_loss)

        # Computing margin of the contract
        margin = round(margin + gain_or_loss, 2)
        i.append(margin)

        # Checking if margin call is necessary
        if margin < maintenance_margin:
            margin_call = round(og_margin - margin, 2)
            margin = og_margin
        else:
            margin_call = '-'

        i.append(margin_call)
        position += 1

    df = pd.DataFrame(data[1:], columns=data[0])
    df[' '] = ""
    df['Variables Used'] = ""
    df['Values'] = ""

    # Initialisation and Declaration
    df.loc[0, 'Variables Used'] = 'VaR Margin'
    df.loc[1, 'Variables Used'] = 'Applicable Margin Rate'
    df.loc[2, 'Variables Used'] = 'Lot Size'
    df.loc[3, 'Variables Used'] = 'Initial Margin'
    df.loc[4, 'Variables Used'] = 'Maintenance Margin'
    df.loc[6, 'Variables Used'] = 'Reference for the data:'
    df.loc[7, 'Variables Used'] = link

    # Importing data from generated sheet
    df.loc[0, 'Values'] = VaR_Margin
    df.loc[1, 'Values'] = applicable_margin_rate
    df.loc[2, 'Values'] = lot_size
    df.loc[3, 'Values'] = og_margin
    df.loc[4, 'Values'] = maintenance_margin

    # Checking if the situation is for a short position contract or a long position contract
    pos_type = "Long" if is_long else "Short"
    df.to_excel(f'{pos_type} Position.xlsx', sheet_name=f"{pos_type} Position", index=False)

def main():
    """Main function to handle the overall process."""
    setup_input_file()

    df = pd.read_excel('FSD_Input.xlsx', sheet_name="Variables", header=None, index_col=0)
    variables = df.values.tolist()

    # Globalisation, Initialisation and Declaration
    global VaR_Margin, applicable_margin_rate, lot_size, settle_price, link, margin, maintenance_margin, og_margin, data
    VaR_Margin = variables[0][0]
    applicable_margin_rate = float(variables[1][0])
    lot_size = variables[2][0]
    settle_price = float(variables[3][0])
    link = str(variables[4][0])

    # Computing essential values
    margin = round(float(lot_size * applicable_margin_rate * settle_price), 2)
    maintenance_margin = round(float(lot_size * VaR_Margin * settle_price), 2)
    og_margin = margin

    df = pd.read_excel('FSD_input.xlsx', sheet_name="Computing Data")
    headers = df.columns.tolist()
    data = [
        [element for element in row if not pd.isna(element)]
        for row in df.values.tolist()
    ]
    data.insert(0, headers)

    compute_position(data, 0, margin, lot_size, is_long=True)

    for index in range(1, len(data)):
        data[index] = data[index][:-3]

    compute_position(data, 0, margin, lot_size, is_long=False)
    
    merge_workbooks('Long Position.xlsx', 'Short Position.xlsx', 'FSD_Output.xlsx')
    print('All tasks done successfully! Find the file "FSD_Output.xlsx" for the computed details.')

if __name__ == "__main__":
    main()